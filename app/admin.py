"""
Admin-Blueprint (für Administratoren und Superadministratoren)
"""
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_from_directory, send_file, make_response
from flask_login import login_required, current_user
from app import db
from app.models import User, TrainerProfile, Certificate, CoachingExperience
from app.forms import TrainerProfileForm, CertificateForm, CoachingExperienceForm, ChangeRoleForm, EditUserForm
from app.utils import save_uploaded_file, delete_file
from app.backup_utils import create_full_backup
from config import Config
from pathlib import Path
import os

bp = Blueprint('admin', __name__)

def require_admin():
    """Prüft ob Benutzer Trainer verwalten kann (Administrator oder Superadministrator)"""
    if not current_user.can_manage_trainers():
        flash('Zugriff verweigert. Diese Seite ist nur für Administratoren verfügbar.', 'error')
        return redirect(url_for('main.dashboard'))
    return None

def require_superadmin():
    """Prüft ob Benutzer Superadministrator ist"""
    if not current_user.is_superadmin():
        flash('Zugriff verweigert. Diese Funktion ist nur für Superadministratoren verfügbar.', 'error')
        return redirect(url_for('admin.overview'))
    return None

@bp.route('/')
@login_required
def overview():
    """Admin-Übersicht"""
    result = require_admin()
    if result:
        return result
    
    # Statistiken
    total_users = User.query.count()
    # Nur aktive Trainer zählen/anzeigen
    total_trainers = TrainerProfile.query.filter(TrainerProfile.active == True).count()
    total_certificates = Certificate.query.count()
    
    # Kürzlich hinzugefügte aktive Trainer
    recent_trainers = TrainerProfile.query.filter(TrainerProfile.active == True).order_by(
        TrainerProfile.created_at.desc()
    ).limit(10).all()
    
    # Gesamtübersicht aller aktiven Trainer
    all_trainers = TrainerProfile.query.filter(TrainerProfile.active == True).order_by(
        TrainerProfile.last_name,
        TrainerProfile.first_name
    ).all()
    
    return render_template(
        'admin/overview.html',
        total_users=total_users,
        total_trainers=total_trainers,
        total_certificates=total_certificates,
        recent_trainers=recent_trainers,
        all_trainers=all_trainers
    )

@bp.route('/users', strict_slashes=False)
@bp.route('/users/')
@login_required
def users():
    """Benutzer-Übersicht (nur für Superadministratoren)"""
    result = require_superadmin()
    if result:
        return result
    
    users_list = User.query.order_by(User.created_at.desc()).all()
    return render_template('admin/users.html', users=users_list)

@bp.route('/users/<int:user_id>', methods=['GET', 'POST'])
@login_required
def user_detail(user_id):
    """Benutzer-Details mit Rollenbearbeitung (nur für Superadministratoren)"""
    result = require_superadmin()
    if result:
        return result
    
    # Lade User frisch aus der Datenbank
    user = User.query.get_or_404(user_id)
    # Expire nur diesen User, um sicherzustellen, dass wir frische Daten bekommen
    db.session.expire(user)
    form = EditUserForm()
    
    # Verhindere Selbständerung der Rolle
    can_edit_roles = user.id != current_user.id
    
    # Verarbeite Formular-Submit
    if request.method == 'POST' and can_edit_roles:
        if 'update_roles' in request.form:
            old_role_name = user.get_role_name()
            
            # Checkboxen senden keinen Wert, wenn sie nicht angehakt sind
            # Daher prüfen wir direkt request.form statt form.data
            is_superadmin = 'is_superadmin' in request.form
            is_admin = 'is_admin' in request.form
            is_coach = 'is_coach' in request.form
            
            # Mindestens eine Rolle muss aktiviert sein
            if not (is_superadmin or is_admin or is_coach):
                flash('Mindestens eine Rolle muss aktiviert sein.', 'error')
            else:
                # Verwende direkte SQL-Abfrage um Namenskonflikte zu vermeiden und sicherzustellen, dass Änderungen gespeichert werden
                from sqlalchemy import text
                db.session.execute(
                    text("""
                        UPDATE users 
                        SET is_superadmin = :is_superadmin, 
                            is_admin = :is_admin, 
                            is_coach = :is_coach 
                        WHERE id = :user_id
                    """),
                    {
                        'is_superadmin': int(is_superadmin),
                        'is_admin': int(is_admin),
                        'is_coach': int(is_coach),
                        'user_id': user.id
                    }
                )
                db.session.commit()
                
                flash(f'Die Rollen von Benutzer {user.username} wurden erfolgreich geändert.', 'success')
                return redirect(url_for('admin.user_detail', user_id=user.id))
    
    # Lade User nochmal frisch aus der DB und setze Formular-Werte
    # Verwende expire und refresh um sicherzustellen, dass wir frische Daten haben
    db.session.expire(user)
    db.session.refresh(user)
    
    # Lade die Rollenwerte direkt aus der Datenbank über eine neue Query
    # Das umgeht alle Caching-Probleme und Namenskonflikte
    from sqlalchemy import text
    result = db.session.execute(
        text("SELECT is_superadmin, is_admin, is_coach FROM users WHERE id = :user_id"),
        {"user_id": user.id}
    ).first()
    
    if result:
        form.is_superadmin.data = bool(result[0]) if result[0] is not None else False
        form.is_admin.data = bool(result[1]) if result[1] is not None else False
        form.is_coach.data = bool(result[2]) if result[2] is not None else False
    else:
        # Fallback falls Query fehlschlägt
        form.is_superadmin.data = False
        form.is_admin.data = False
        form.is_coach.data = False
    
    return render_template('admin/user_detail.html', user=user, form=form, can_edit_roles=can_edit_roles)

@bp.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
def delete_user(user_id):
    """Benutzer deaktivieren (nur für Superadministratoren) - setzt active=False"""
    result = require_superadmin()
    if result:
        return result
    
    user = User.query.get_or_404(user_id)
    
    # Verhindere Selbstdeaktivierung
    if user.id == current_user.id:
        flash('Sie können sich nicht selbst deaktivieren.', 'error')
        return redirect(url_for('admin.users'))
    
    # Deaktiviere Benutzer (kann sich nicht mehr anmelden)
    user.active = False
    db.session.commit()
    
    flash(f'Benutzer {user.username} wurde erfolgreich deaktiviert.', 'success')
    return redirect(url_for('admin.users'))

@bp.route('/trainers')
@login_required
def trainers():
    """Trainer-Übersicht - zeigt standardmäßig nur aktive Trainer"""
    result = require_admin()
    if result:
        return result
    
    # Zeige nur aktive Trainer (für alle)
    trainers_list = TrainerProfile.query.filter(TrainerProfile.active == True).order_by(
        TrainerProfile.last_name, TrainerProfile.first_name
    ).all()
    return render_template('admin/trainers.html', trainers=trainers_list, show_inactive=False)

@bp.route('/trainers/inactive')
@login_required
def trainers_inactive():
    """Trainer-Übersicht - zeigt nur inaktive Trainer"""
    result = require_admin()
    if result:
        return result
    
    # Zeige nur inaktive Trainer
    trainers_list = TrainerProfile.query.filter(TrainerProfile.active == False).order_by(
        TrainerProfile.last_name, TrainerProfile.first_name
    ).all()
    return render_template('admin/trainers.html', trainers=trainers_list, show_inactive=True)

@bp.route('/coaches')
@login_required
def coaches_list():
    """Coach-Liste mit Filterung und Sortierung - zeigt nur Trainer mit Rolle 'coach'"""
    result = require_admin()
    if result:
        return result
    
    # Filter- und Sortierparameter aus Query-String lesen
    team_filter = request.args.get('team', '')
    sort_by = request.args.get('sort', 'name')  # name, coaching_years, certificates
    
    # Basis-Query - hole alle TrainerProfile von Benutzern mit is_coach=True
    # Verwende direkte SQL-Abfrage um sicherzustellen, dass alle erfasst werden
    from sqlalchemy import text
    
    # Hole alle user_ids mit is_coach=True direkt aus der DB
    coach_user_ids = db.session.execute(
        text("SELECT id FROM users WHERE is_coach = 1")
    ).fetchall()
    coach_user_ids = [row[0] for row in coach_user_ids]
    
    if not coach_user_ids:
        coaches = []
    else:
        # Lade alle TrainerProfile für diese Benutzer, aber nur wenn sie aktiv sind
        coaches = TrainerProfile.query.filter(
            TrainerProfile.user_id.in_(coach_user_ids),
            TrainerProfile.active == True
        ).all()
        
        # Team-Filter anwenden (wenn ein Team ausgewählt wurde)
        if team_filter:
            coaches = [c for c in coaches if any(
                exp.team == team_filter for exp in c.coaching_experiences
            )]
        
        # Sortierung anwenden
        if sort_by == 'name':
            coaches.sort(key=lambda c: (c.last_name or '', c.first_name or ''))
        elif sort_by == 'coaching_years':
            coaches.sort(key=lambda c: c.get_total_coaching_years(), reverse=True)
        elif sort_by == 'certificates':
            coaches.sort(key=lambda c: c.certificates.count(), reverse=True)
    
    # Verfügbare Teams für Filter
    available_teams = CoachingExperience.TEAMS
    
    return render_template(
        'admin/coaches_list.html',
        coaches=coaches,
        team_filter=team_filter,
        sort_by=sort_by,
        available_teams=available_teams
    )

@bp.route('/coaches/export')
@login_required
def coaches_export():
    """Coach-Liste als Excel exportieren"""
    result = require_admin()
    if result:
        return result
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from datetime import datetime
    except ImportError:
        flash('Excel-Export benötigt das Paket "openpyxl". Bitte installieren Sie es mit: pip install openpyxl', 'error')
        return redirect(url_for('admin.coaches_list'))
    
    # Filter- und Sortierparameter aus Query-String lesen (wie in coaches_list)
    team_filter = request.args.get('team', '')
    sort_by = request.args.get('sort', 'name')
    
    # Basis-Query - hole alle TrainerProfile von Benutzern mit is_coach=True
    # Verwende direkte SQL-Abfrage um sicherzustellen, dass alle erfasst werden
    from sqlalchemy import text
    
    # Hole alle user_ids mit is_coach=True direkt aus der DB
    coach_user_ids = db.session.execute(
        text("SELECT id FROM users WHERE is_coach = 1")
    ).fetchall()
    coach_user_ids = [row[0] for row in coach_user_ids]
    
    if not coach_user_ids:
        coaches = []
    else:
        # Lade alle TrainerProfile für diese Benutzer, aber nur wenn sie aktiv sind
        coaches = TrainerProfile.query.filter(
            TrainerProfile.user_id.in_(coach_user_ids),
            TrainerProfile.active == True
        ).all()
        
        # Team-Filter anwenden (wenn ein Team ausgewählt wurde)
        if team_filter:
            coaches = [c for c in coaches if any(
                exp.team == team_filter for exp in c.coaching_experiences
            )]
        
        # Sortierung anwenden
        if sort_by == 'name':
            coaches.sort(key=lambda c: (c.last_name or '', c.first_name or ''))
        elif sort_by == 'coaching_years':
            coaches.sort(key=lambda c: c.get_total_coaching_years(), reverse=True)
        elif sort_by == 'certificates':
            coaches.sort(key=lambda c: c.certificates.count(), reverse=True)
    
    # Excel-Workbook erstellen
    wb = Workbook()
    ws = wb.active
    ws.title = "Coach-Liste"
    
    # Header-Formatierung
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Header-Zeile
    headers = [
        'Name', 'Vorname', 'Nachname', 'Lizenznummer', 'Aktuelle Funktion',
        'E-Mail privat', 'E-Mail geschäftlich', 'Mobiltelefon privat', 'Mobiltelefon geschäftlich',
        'Strasse', 'PLZ', 'Ort',
        'Coaching-Jahre', 'Anzahl Zertifikate'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Daten einfügen
    for row_num, coach in enumerate(coaches, 2):
        # Teams sammeln (einzigartig)
        unique_teams = set()
        for exp in coach.coaching_experiences.all():
            unique_teams.add(exp.team)
        teams_str = ', '.join(sorted(unique_teams)) if unique_teams else ''
        
        ws.cell(row=row_num, column=1, value=f"{coach.first_name} {coach.last_name}")
        ws.cell(row=row_num, column=2, value=coach.first_name)
        ws.cell(row=row_num, column=3, value=coach.last_name)
        ws.cell(row=row_num, column=4, value=coach.license_number or '')
        ws.cell(row=row_num, column=5, value=coach.function_club or '')
        ws.cell(row=row_num, column=6, value=coach.email_private or '')
        ws.cell(row=row_num, column=7, value=coach.email_business or '')
        ws.cell(row=row_num, column=8, value=coach.phone_private or '')
        ws.cell(row=row_num, column=9, value=coach.phone_business or '')
        ws.cell(row=row_num, column=10, value=coach.street or '')
        ws.cell(row=row_num, column=11, value=coach.postal_code or '')
        ws.cell(row=row_num, column=12, value=coach.city or '')
        ws.cell(row=row_num, column=13, value=coach.get_total_coaching_years())
        ws.cell(row=row_num, column=14, value=coach.certificates.count())
    
    # Spaltenbreite anpassen
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20
    
    # Header-Zeile einfrieren
    ws.freeze_panes = 'A2'
    
    # Dateiname generieren
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'coaches_export_{timestamp}.xlsx'
    
    # Excel-Datei in Memory speichern
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Response erstellen
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

@bp.route('/trainers/<int:trainer_id>')
@login_required
def trainer_detail(trainer_id):
    """Trainer-Details anzeigen"""
    result = require_admin()
    if result:
        return result
    
    trainer = TrainerProfile.query.get_or_404(trainer_id)
    
    # Lade active-Status direkt aus der DB um sicherzustellen, dass wir den korrekten Wert haben
    from sqlalchemy import text
    active_value = db.session.execute(
        text("SELECT active FROM trainer_profiles WHERE id = :trainer_id"),
        {"trainer_id": trainer_id}
    ).scalar()
    
    # Setze active-Status explizit (umgeht SQLAlchemy-Caching)
    trainer.active = bool(active_value) if active_value is not None else True
    
    # Refresh Trainer-Objekt um sicherzustellen, dass wir aktuelle Daten haben
    db.session.expire(trainer)
    db.session.refresh(trainer)
    
    # Lade Zertifikate (absteigend sortiert)
    from sqlalchemy import desc
    certificates = trainer.certificates.order_by(
        desc(Certificate.issue_date),
        desc(Certificate.created_at)
    ).all()
    
    # Lade Coaching-Erfahrungen (absteigend sortiert, laufende zuerst)
    from sqlalchemy import case
    coaching_experiences = trainer.coaching_experiences.order_by(
        case((CoachingExperience.end_year.is_(None), 0), else_=1),  # Laufende (NULL) zuerst
        desc(CoachingExperience.end_year),
        desc(CoachingExperience.start_year)
    ).all()
    
    return render_template('admin/trainer_detail.html', trainer=trainer, certificates=certificates, coaching_experiences=coaching_experiences)

@bp.route('/trainers/<int:trainer_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_trainer(trainer_id):
    """Trainer-Profil bearbeiten"""
    result = require_admin()
    if result:
        return result
    
    trainer = TrainerProfile.query.get_or_404(trainer_id)
    form = TrainerProfileForm(obj=trainer)
    
    if form.validate_on_submit():
        # Profilbild wird separat behandelt (über separaten Endpoint)
        profile_image_data = form.profile_image.data
        form.profile_image.data = None
        form.populate_obj(trainer)
        form.profile_image.data = profile_image_data
        
        db.session.commit()
        flash('Trainer-Profil wurde erfolgreich aktualisiert.', 'success')
        return redirect(url_for('admin.trainer_detail', trainer_id=trainer.id))
    elif request.method == 'POST':
        # Formular wurde abgesendet, aber nicht validiert - zeige Fehler
        flash('Bitte korrigieren Sie die Fehler im Formular.', 'error')
    
    # Lade Coaching-Erfahrungen (absteigend sortiert, laufende zuerst)
    from sqlalchemy import case, desc
    coaching_experiences = trainer.coaching_experiences.order_by(
        case((CoachingExperience.end_year.is_(None), 0), else_=1),  # Laufende (NULL) zuerst
        desc(CoachingExperience.end_year),
        desc(CoachingExperience.start_year)
    ).all()
    
    # Lade Zertifikate (absteigend sortiert nach Ausstellungsdatum)
    certificates = trainer.certificates.order_by(
        desc(Certificate.issue_date),
        desc(Certificate.created_at)
    ).all()
    
    return render_template('admin/edit_trainer.html', form=form, trainer=trainer, coaching_experiences=coaching_experiences, certificates=certificates)

@bp.route('/trainers/<int:trainer_id>/delete', methods=['POST'])
@login_required
def delete_trainer(trainer_id):
    """Trainer deaktivieren (nur für Superadministratoren) - setzt active=False"""
    result = require_superadmin()
    if result:
        return result
    
    trainer = TrainerProfile.query.get_or_404(trainer_id)
    user = trainer.user
    
    # Verhindere Selbstdeaktivierung
    if user.id == current_user.id:
        flash('Sie können sich nicht selbst deaktivieren.', 'error')
        return redirect(url_for('admin.trainers'))
    
    # Deaktiviere Trainer (wird dadurch nicht mehr in Listen angezeigt)
    trainer.active = False
    db.session.commit()
    
    flash(f'Trainer {trainer.first_name} {trainer.last_name} wurde erfolgreich deaktiviert.', 'success')
    return redirect(url_for('admin.trainers'))

@bp.route('/trainers/<int:trainer_id>/toggle-status', methods=['POST'])
@login_required
def toggle_trainer_status(trainer_id):
    """Trainer aktivieren/deaktivieren (nur für Superadministratoren)"""
    result = require_superadmin()
    if result:
        return result
    
    trainer = TrainerProfile.query.get_or_404(trainer_id)
    user = trainer.user
    
    # Verhindere Selbständerung des Status
    if user.id == current_user.id:
        flash('Sie können Ihren eigenen Status nicht ändern.', 'error')
        return redirect(url_for('admin.trainer_detail', trainer_id=trainer.id))
    
    # Lade aktuellen Status direkt aus DB
    from sqlalchemy import text
    current_active = db.session.execute(
        text("SELECT active FROM trainer_profiles WHERE id = :trainer_id"),
        {"trainer_id": trainer_id}
    ).scalar()
    
    # Toggle Status
    new_status = 1 if not bool(current_active) else 0
    
    # Update direkt in DB für Zuverlässigkeit
    db.session.execute(
        text("UPDATE trainer_profiles SET active = :active WHERE id = :trainer_id"),
        {"active": new_status, "trainer_id": trainer_id}
    )
    db.session.commit()
    
    status_text = 'aktiviert' if new_status == 1 else 'deaktiviert'
    flash(f'Trainer {trainer.first_name} {trainer.last_name} wurde erfolgreich {status_text}.', 'success')
    return redirect(url_for('admin.trainer_detail', trainer_id=trainer.id))

@bp.route('/trainers/<int:trainer_id>/upload-image', methods=['POST'])
@login_required
def upload_profile_image(trainer_id):
    """Profilbild-Upload (separater Endpoint für Admin)"""
    result = require_admin()
    if result:
        return result
    
    trainer = TrainerProfile.query.get_or_404(trainer_id)
    
    if 'profile_image' not in request.files:
        flash('Keine Datei ausgewählt.', 'error')
        return redirect(url_for('admin.edit_trainer', trainer_id=trainer.id))
    
    file = request.files['profile_image']
    
    if file.filename == '':
        flash('Keine Datei ausgewählt.', 'error')
        return redirect(url_for('admin.edit_trainer', trainer_id=trainer.id))
    
    # Lösche altes Profilbild falls vorhanden
    if trainer.profile_image_path:
        delete_file(trainer.profile_image_path)
    
    # Speichere neues Profilbild
    image_path = save_uploaded_file(
        file,
        'PROFILE_IMAGES',
        prefix=f'profile_{trainer.user_id}'
    )
    
    if image_path:
        trainer.profile_image_path = image_path
        db.session.commit()
        flash('Profilbild wurde erfolgreich aktualisiert.', 'success')
    else:
        flash('Fehler beim Hochladen des Profilbildes.', 'error')
    
    return redirect(url_for('admin.edit_trainer', trainer_id=trainer.id))

@bp.route('/trainers/<int:trainer_id>/certificates/add', methods=['GET', 'POST'])
@login_required
def add_certificate(trainer_id):
    """Zertifikat für Trainer hinzufügen (Admin)"""
    result = require_admin()
    if result:
        return result
    
    trainer = TrainerProfile.query.get_or_404(trainer_id)
    form = CertificateForm()
    
    if form.validate_on_submit():
        file_path = None
        if form.file.data:
            file_path = save_uploaded_file(
                form.file.data,
                'CERTIFICATES',
                prefix=f'cert_{trainer.user_id}'
            )
        
        if file_path or not form.file.data:  # Erlaube Zertifikat ohne Datei
            certificate = Certificate(
                trainer_profile_id=trainer.id,
                title=form.title.data,
                issuer=form.issuer.data,
                issue_date=form.issue_date.data,
                expiry_date=form.expiry_date.data,
                description=form.description.data,
                training_type=form.training_type.data if form.training_type.data else None,
                file_path=file_path,
                file_type=form.file.data.filename.rsplit('.', 1)[1].lower() if form.file.data and '.' in form.file.data.filename else None
            )
            db.session.add(certificate)
            db.session.commit()
            flash('Zertifikat wurde erfolgreich hinzugefügt.', 'success')
            return redirect(url_for('admin.trainer_detail', trainer_id=trainer.id))
        else:
            flash('Fehler beim Hochladen der Datei.', 'error')
    
    return render_template('admin/add_certificate.html', form=form, trainer=trainer)

@bp.route('/certificates/<int:cert_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_certificate(cert_id):
    """Zertifikat bearbeiten (Admin)"""
    result = require_admin()
    if result:
        return result
    
    certificate = Certificate.query.get_or_404(cert_id)
    trainer_id = certificate.trainer_profile_id
    
    form = CertificateForm(obj=certificate)
    
    # Datei-Feld optional machen für Edit
    from wtforms.validators import Optional
    from flask_wtf.file import FileAllowed
    form.file.validators = [Optional(), FileAllowed(['pdf', 'png', 'jpg', 'jpeg', 'gif'], 'Nur PDF und Bilddateien erlaubt')]
    
    if form.validate_on_submit():
        # Wenn eine neue Datei hochgeladen wurde
        if form.file.data:
            # Lösche alte Datei
            delete_file(certificate.file_path)
            
            # Speichere neue Datei
            trainer = TrainerProfile.query.get_or_404(trainer_id)
            file_path = save_uploaded_file(
                form.file.data,
                'CERTIFICATES',
                prefix=f'cert_{trainer.user_id}'
            )
            if file_path:
                certificate.file_path = file_path
                certificate.file_type = form.file.data.filename.rsplit('.', 1)[1].lower()
        
        # Aktualisiere andere Felder
        certificate.title = form.title.data
        certificate.issuer = form.issuer.data
        certificate.issue_date = form.issue_date.data
        certificate.expiry_date = form.expiry_date.data
        certificate.description = form.description.data
        certificate.training_type = form.training_type.data if form.training_type.data else None
        
        db.session.commit()
        flash('Zertifikat wurde erfolgreich aktualisiert.', 'success')
        return redirect(url_for('admin.trainer_detail', trainer_id=trainer_id))
    
    trainer = TrainerProfile.query.get_or_404(trainer_id)
    return render_template('admin/edit_certificate.html', form=form, certificate=certificate, trainer=trainer)

@bp.route('/certificates/<int:cert_id>/delete', methods=['POST'])
@login_required
def delete_certificate(cert_id):
    """Zertifikat löschen (nur für Superadministratoren)"""
    result = require_superadmin()
    if result:
        return result
    
    certificate = Certificate.query.get_or_404(cert_id)
    
    # Lösche Datei
    if certificate.file_path:
        delete_file(certificate.file_path)
    
    trainer_id = certificate.trainer_profile_id
    
    # Lösche Zertifikat
    db.session.delete(certificate)
    db.session.commit()
    
    flash('Zertifikat wurde erfolgreich gelöscht.', 'success')
    return redirect(url_for('admin.trainer_detail', trainer_id=trainer_id))

@bp.route('/trainers/<int:trainer_id>/cv/add', methods=['GET', 'POST'])
@login_required
def add_coaching_experience(trainer_id):
    """Neue Coaching-Erfahrung für Trainer hinzufügen (Admin)"""
    result = require_admin()
    if result:
        return result
    
    trainer = TrainerProfile.query.get_or_404(trainer_id)
    form = CoachingExperienceForm()
    
    if form.validate_on_submit():
        experience = CoachingExperience(
            trainer_profile_id=trainer.id,
            start_year=form.start_year.data,
            end_year=form.end_year.data,
            coaching_role=form.coaching_role.data,
            team=form.team.data
        )
        db.session.add(experience)
        db.session.commit()
        flash('Coaching-Erfahrung wurde erfolgreich hinzugefügt.', 'success')
        return redirect(url_for('admin.trainer_detail', trainer_id=trainer.id))
    
    return render_template('admin/add_coaching_experience.html', form=form, trainer=trainer)

@bp.route('/cv/<int:exp_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_coaching_experience(exp_id):
    """Coaching-Erfahrung bearbeiten (Admin)"""
    result = require_admin()
    if result:
        return result
    
    experience = CoachingExperience.query.get_or_404(exp_id)
    trainer_id = experience.trainer_profile_id
    
    form = CoachingExperienceForm(obj=experience)
    
    if form.validate_on_submit():
        form.populate_obj(experience)
        db.session.commit()
        flash('Coaching-Erfahrung wurde erfolgreich aktualisiert.', 'success')
        return redirect(url_for('admin.trainer_detail', trainer_id=trainer_id))
    
    trainer = TrainerProfile.query.get_or_404(trainer_id)
    return render_template('admin/edit_coaching_experience.html', form=form, experience=experience, trainer=trainer)

@bp.route('/cv/<int:exp_id>/delete', methods=['POST'])
@login_required
def delete_coaching_experience(exp_id):
    """Coaching-Erfahrung löschen (nur für Superadministratoren)"""
    result = require_superadmin()
    if result:
        return result
    
    experience = CoachingExperience.query.get_or_404(exp_id)
    trainer_id = experience.trainer_profile_id
    
    db.session.delete(experience)
    db.session.commit()
    
    flash('Coaching-Erfahrung wurde erfolgreich gelöscht.', 'success')
    return redirect(url_for('admin.trainer_detail', trainer_id=trainer_id))

@bp.route('/users/<int:user_id>/change-password', methods=['GET', 'POST'])
@login_required
def change_user_password(user_id):
    """Passwort eines Benutzers ändern (nur für Superadministratoren)"""
    result = require_superadmin()
    if result:
        return result
    
    user = User.query.get_or_404(user_id)
    
    # Superadmins können nur von anderen Superadmins bearbeitet werden
    if user.is_superadmin() and not current_user.is_superadmin():
        flash('Das Passwort von Superadministratoren kann nur von anderen Superadministratoren geändert werden.', 'error')
        return redirect(url_for('admin.users'))
    
    from app.forms import ChangePasswordForm
    form = ChangePasswordForm()
    
    if form.validate_on_submit():
        user.set_password(form.new_password.data)
        db.session.commit()
        flash(f'Das Passwort für Benutzer {user.username} wurde erfolgreich geändert.', 'success')
        return redirect(url_for('admin.user_detail', user_id=user.id))
    
    return render_template('admin/change_password.html', form=form, user=user)

@bp.route('/users/<int:user_id>/change-role', methods=['GET', 'POST'])
@login_required
def change_user_role(user_id):
    """Rolle eines Benutzers ändern (nur für Superadministratoren)"""
    result = require_superadmin()
    if result:
        return result
    
    user = User.query.get_or_404(user_id)
    
    # Verhindere Selbständerung der Rolle
    if user.id == current_user.id:
        flash('Sie können Ihre eigene Rolle nicht ändern.', 'error')
        return redirect(url_for('admin.user_detail', user_id=user.id))
    
    form = ChangeRoleForm()
    
    # Setze aktuelle Rolle als Standard
    if request.method == 'GET':
        if user.is_superadmin():
            form.role.data = 'superadmin'
        elif user.is_admin():
            form.role.data = 'admin'
        elif user.is_coach():
            form.role.data = 'coach'
    
    if form.validate_on_submit():
        old_role_name = user.get_role_name()
        new_role_name = form.role.data
        
        # Setze alle Flags auf False zuerst - direkte Zuweisung für SQLAlchemy
        User.query.filter_by(id=user.id).update({
            'is_superadmin': False,
            'is_admin': False,
            'is_coach': False
        })
        
        # Setze die gewählte Rolle direkt in der Datenbank
        if new_role_name == 'superadmin':
            User.query.filter_by(id=user.id).update({'is_superadmin': True})
        elif new_role_name == 'admin':
            User.query.filter_by(id=user.id).update({'is_admin': True})
        elif new_role_name == 'coach':
            User.query.filter_by(id=user.id).update({'is_coach': True})
        
        db.session.commit()
        
        flash(f'Die Rolle von Benutzer {user.username} wurde von "{old_role_name}" zu "{new_role_name}" geändert.', 'success')
        return redirect(url_for('admin.user_detail', user_id=user.id))
    
    return render_template('admin/change_role.html', form=form, user=user)

@bp.route('/users/<int:user_id>/toggle-status', methods=['POST'])
@login_required
def toggle_user_status(user_id):
    """Benutzer aktivieren/deaktivieren (nur für Superadministratoren)"""
    result = require_superadmin()
    if result:
        return result
    
    user = User.query.get_or_404(user_id)
    
    # Verhindere Selbständerung des Status
    if user.id == current_user.id:
        flash('Sie können Ihren eigenen Status nicht ändern.', 'error')
        return redirect(url_for('admin.user_detail', user_id=user.id))
    
    # Toggle Status
    user.active = not user.active
    db.session.commit()
    
    status_text = 'aktiviert' if user.active else 'deaktiviert'
    flash(f'Benutzer {user.username} wurde erfolgreich {status_text}.', 'success')
    return redirect(url_for('admin.user_detail', user_id=user.id))

@bp.route('/backup', methods=['GET', 'POST'])
@login_required
def backup():
    """Backup erstellen (nur für Superadministratoren)"""
    result = require_superadmin()
    if result:
        return result
    
    if request.method == 'POST':
        # Erstelle Backup
        success, backup_path, error_msg = create_full_backup()
        
        if success:
            flash('Backup wurde erfolgreich erstellt!', 'success')
            # Sende Datei zum Download
            return send_file(
                backup_path,
                as_attachment=True,
                download_name=os.path.basename(backup_path),
                mimetype='application/zip'
            )
        else:
            flash(f'Fehler beim Erstellen des Backups: {error_msg}', 'error')
    
    return render_template('admin/backup.html')

@bp.route('/files/<path:filename>')
@login_required
def serve_file(filename):
    """Serviert hochgeladene Dateien (für Administratoren)"""
    result = require_admin()
    if result:
        return result
    
    # Normalisiere Dateiname (konvertiere Backslashes zu Forward-Slashes)
    filename_normalized = filename.replace('\\', '/')
    file_path = Path(Config.UPLOAD_FOLDER) / filename_normalized
    
    if file_path.exists():
        upload_folder = Path(Config.UPLOAD_FOLDER)
        return send_from_directory(str(upload_folder), filename_normalized)
    
    flash('Datei nicht gefunden.', 'error')
    return redirect(url_for('admin.overview'))

